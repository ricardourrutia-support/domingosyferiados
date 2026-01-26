import io
import re
from datetime import datetime, date
from typing import List, Set, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Font  # <--- Importante para las negritas

# ----------------------------
# Config
# ----------------------------
st.set_page_config(page_title="Reporte Domingos y Feriados", layout="wide")

INVALID_TURNOS = {"L", ""}  # "L" y vacío NO cuentan como trabajado. COON1 sí cuenta.

# ----------------------------
# Helpers
# ----------------------------
def _normalize_turno(x) -> str:
    """Normaliza el turno a string limpio. None/NaN -> ''."""
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    s = str(x).strip()
    # Si viene como "nan" texto
    if s.lower() == "nan" or s.lower() == "none":
        return ""
    return s


def is_turno_valido(turno: str) -> bool:
    """Valida turno: todo cuenta (incluido COON1) excepto 'L' y vacío."""
    t = _normalize_turno(turno)
    return t not in INVALID_TURNOS


def detect_date_columns(df: pd.DataFrame, meta_cols: List[str]) -> List:
    """
    Detecta columnas de fecha del reporte.
    Maneja encabezados que vienen como datetime real o como string 'dd-mm-yyyy'.
    """
    date_cols = []
    for c in df.columns:
        if c in meta_cols:
            continue
        
        # Caso 1: Excel ya lo leyó como objeto fecha/timestamp
        if isinstance(c, (pd.Timestamp, datetime, date)):
            date_cols.append(c)
            continue
        
        # Caso 2: Viene como string "01-01-2026"
        if isinstance(c, str):
            c2 = c.strip()
            try:
                # IMPORTANTE: dayfirst=True para formato Latino (dd-mm-yyyy)
                _ = pd.to_datetime(c2, errors="raise", dayfirst=True)
                date_cols.append(c)
            except Exception:
                # Si falla, asumimos que no es fecha (ej: columna de totales del excel)
                pass
    return date_cols


def parse_holidays(text: str) -> Set[pd.Timestamp]:
    """
    Acepta fechas manuales separadas por coma o salto de línea.
    """
    if not text or not text.strip():
        return set()

    parts = re.split(r"[,\n;]+", text.strip())
    holidays = set()

    for p in parts:
        p = p.strip()
        if not p:
            continue

        # Normaliza separadores
        p_norm = p.replace(".", "-").replace("/", "-")

        # Intento con dayfirst=True (Chile/Latam)
        dt_val = pd.to_datetime(p_norm, errors="coerce", dayfirst=True)
        
        if pd.isna(dt_val):
             # Intento alternativo formato ISO
            dt_val = pd.to_datetime(p_norm, errors="coerce", dayfirst=False)

        if pd.isna(dt_val):
            st.warning(f"⚠️ No se pudo interpretar la fecha de feriado manual: '{p}'. Se omitirá.")
            continue

        holidays.add(pd.Timestamp(dt_val.date()))

    return holidays


def build_summary(
    df: pd.DataFrame,
    meta_cols: List[str],
    date_cols: List,
    holidays: Set[pd.Timestamp],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, str]:
    """
    Procesa la data para generar los reportes.
    """
    if not date_cols:
        raise ValueError("No se detectaron columnas de fecha en el archivo. Verifica que los encabezados tengan formato de fecha (ej: 01-01-2026).")

    # Melt largo: Transforma de formato ancho a formato largo
    dfl = df.melt(id_vars=meta_cols, value_vars=date_cols, var_name="Fecha_Header", value_name="Turno")
    dfl["Turno_norm"] = dfl["Turno"].apply(_normalize_turno)
    dfl["Trabajado"] = dfl["Turno_norm"].apply(is_turno_valido)

    # Convertir Header a Fecha real
    # Usamos dayfirst=True para asegurar que '01-02' sea 1 de Febrero y no 2 de Enero
    dfl["Fecha_dt"] = pd.to_datetime(dfl["Fecha_Header"], errors="coerce", dayfirst=True)
    
    if dfl["Fecha_dt"].isna().all():
        raise ValueError("Error convirtiendo columnas a fecha. Revisa el formato de los encabezados.")

    dfl["Fecha_dt"] = dfl["Fecha_dt"].dt.date
    dfl["Fecha_ts"] = dfl["Fecha_dt"].apply(lambda x: pd.Timestamp(x))

    # Definir rango del periodo detectado
    start = min(dfl["Fecha_dt"])
    end = max(dfl["Fecha_dt"])
    periodo_str = f"{start.strftime('%d-%m-%Y')} a {end.strftime('%d-%m-%Y')}"

    # Lógica Domingos (weekday == 6)
    dfl["Es_domingo"] = pd.to_datetime(dfl["Fecha_dt"]).dt.weekday == 6
    dom = dfl[(dfl["Trabajado"]) & (dfl["Es_domingo"])].copy()

    # Lógica Festivos (cruce con input manual)
    if holidays:
        fest = dfl[(dfl["Trabajado"]) & (dfl["Fecha_ts"].isin(holidays))].copy()
    else:
        fest = dfl.iloc[0:0].copy()

    # Función de agregación
    def agg_table(sub: pd.DataFrame, label_count: str, label_dates: str) -> pd.DataFrame:
        if sub.empty:
            base = df[meta_cols].drop_duplicates().copy()
            base[label_count] = 0
            base[label_dates] = ""
            return base

        grp = (
            sub.groupby(meta_cols, dropna=False)["Fecha_dt"]
            .apply(lambda s: sorted(set(s)))
            .reset_index()
        )
        grp[label_count] = grp["Fecha_dt"].apply(len)
        grp[label_dates] = grp["Fecha_dt"].apply(lambda lst: ", ".join([pd.Timestamp(x).strftime("%d-%m-%Y") for x in lst]))
        grp = grp.drop(columns=["Fecha_dt"])
        return grp

    dom_tbl = agg_table(dom, "Domingos trabajados", "Fechas (domingos)")
    fest_tbl = agg_table(fest, "Festivos trabajados", "Fechas (festivos)")

    # Cruce Final (Total)
    # Reconstruimos la base completa para asegurar que estén todos los colaboradores
    base_users = df[meta_cols].drop_duplicates().reset_index(drop=True)
    
    # Preparamos diccionarios de búsqueda rápida
    # Clave: Tupla de (RUT, Supervisor, etc), Valor: Set de fechas
    def make_lookup(sub_df):
        if sub_df.empty: return {}
        return sub_df.groupby(meta_cols, dropna=False)["Fecha_dt"].apply(set).to_dict()

    dom_map = make_lookup(dom)
    fest_map = make_lookup(fest)

    total_rows = []
    for _, row in base_users.iterrows():
        # Crear clave compuesta basada en las columnas meta
        key = tuple(row[c] for c in meta_cols)
        
        s_dom = dom_map.get(key, set())
        s_fest = fest_map.get(key, set())
        s_all = s_dom | s_fest # Unión de conjuntos (sin duplicados)

        total_rows.append(
            {
                **{c: row[c] for c in meta_cols},
                "Domingos trabajados": len(s_dom),
                "Festivos trabajados": len(s_fest),
                "Total (D + F)": len(s_all),
                "Fechas (domingos)": ", ".join([x.strftime("%d-%m-%Y") for x in sorted(s_dom)]),
                "Fechas (festivos)": ", ".join([x.strftime("%d-%m-%Y") for x in sorted(s_fest)]),
                "Fechas (todas)": ", ".join([x.strftime("%d-%m-%Y") for x in sorted(s_all)]),
            }
        )

    total_tbl = pd.DataFrame(total_rows)

    return dom_tbl, fest_tbl, total_tbl, periodo_str


def export_excel(dom_tbl, fest_tbl, total_tbl, periodo, holidays) -> bytes:
    """Genera el Excel final con formato."""
    output = io.BytesIO()
    feriados_str = ", ".join(sorted([h.strftime("%d-%m-%Y") for h in holidays])) if holidays else "(ninguno)"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        startrow = 4
        
        # Hojas
        sheets = {
            "Domingos": dom_tbl,
            "Festivos": fest_tbl,
            "Resumen Total": total_tbl
        }

        for sheet_name, data in sheets.items():
            data.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow)
            ws = writer.book[sheet_name]
            
            # Encabezado visual
            ws["A1"] = f"Reporte de Asistencias: {sheet_name}"
            # Corrección del error: Usar objeto Font de openpyxl directamente
            ws["A1"].font = Font(bold=True, size=12)
            
            ws["A2"] = f"Periodo: {periodo}"
            ws["A3"] = "Criterio: Se cuentan todos los turnos excepto 'L' y celdas vacías."
            if sheet_name != "Domingos":
                ws["A4"] = f"Feriados considerados: {feriados_str}"
            
            # Ajuste ancho columnas
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col[:200]: # Muestreo primeras 200 filas
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 50)

    return output.getvalue()


# ----------------------------
# UI
# ----------------------------
st.title("📊 Calculadora de Domingos y Feriados Trabajados")

st.markdown("""
Esta herramienta procesa la planilla de turnos y cuenta automáticamente:
1.  **Domingos trabajados** (calculado por calendario).
2.  **Feriados trabajados** (según las fechas que ingreses manualmente).
""")

# Input File
uploaded = st.file_uploader("📂 Cargar archivo (.xlsx o .xls)", type=["xlsx", "xls"])

# Input Feriados
col_f1, col_f2 = st.columns([2, 1])
with col_f1:
    feriados_text = st.text_area(
        "📅 Ingresa los feriados del periodo (separados por coma o enter)",
        placeholder="Ejemplo: 01-01-2026, 15-01-2026",
        height=100
    )
with col_f2:
    st.info("""
    **Reglas de cálculo:**
    - ✅ **Trabajado:** Cualquier texto (ej: AM, PM, COON1).
    - ❌ **No trabajado:** Solo "L" o celda vacía.
    """)

if uploaded:
    try:
        # Detectar motor segun extension
        engine = "openpyxl" if uploaded.name.endswith("xlsx") else "xlrd"
        
        # Leer archivo
        xls_file = pd.ExcelFile(uploaded, engine=engine)
        sheet_name = st.selectbox("Selecciona la hoja con los datos:", xls_file.sheet_names)
        
        df = pd.read_excel(uploaded, sheet_name=sheet_name, engine=engine)

        # Definir columnas de metadatos (Colaborador)
        meta_cols = ["Nombre del Colaborador", "RUT", "Área", "Supervisor"]
        
        # Validación básica de estructura
        missing = [c for c in meta_cols if c not in df.columns]
        if missing:
            st.error(f"❌ Error: No encuentro las columnas: {', '.join(missing)}")
            st.warning("Asegúrate que el Excel tenga las columnas: Nombre del Colaborador, RUT, Área, Supervisor")
            st.stop()

        # Proceso
        date_cols = detect_date_columns(df, meta_cols)
        holidays = parse_holidays(feriados_text)
        
        # Spinner visual
        with st.spinner('Procesando turnos...'):
            dom_tbl, fest_tbl, total_tbl, periodo_str = build_summary(df, meta_cols, date_cols, holidays)

        st.success(f"✅ Procesamiento exitoso. Periodo detectado: **{periodo_str}**")

        # Visualización en pestañas
        tab1, tab2, tab3 = st.tabs(["🔴 Domingos", "🔵 Festivos", "🟢 Resumen Total"])
        
        with tab1:
            st.dataframe(dom_tbl, use_container_width=True, hide_index=True)
        with tab2:
            st.dataframe(fest_tbl, use_container_width=True, hide_index=True)
        with tab3:
            st.dataframe(total_tbl, use_container_width=True, hide_index=True)

        # Descarga
        excel_data = export_excel(dom_tbl, fest_tbl, total_tbl, periodo_str, holidays)
        st.download_button(
            label="⬇️ Descargar Reporte Consolidado (Excel)",
            data=excel_data,
            file_name="Reporte_Domingos_y_Feriados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except ValueError as ve:
        st.error(f"Error de Datos: {ve}")
    except Exception as e:
        st.error("Ocurrió un error inesperado al procesar el archivo.")
        st.exception(e)

else:
    st.info("👆 Carga tu archivo Excel para comenzar.")
